# app.py (Финальная, эталонная версия v22 - Production-Ready)
import json
from copy import deepcopy
from typing import Dict, Any
from pathlib import Path
from io import BytesIO

import streamlit as st
from jsonschema import validate, ValidationError
import pandas as pd
import altair as alt

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment

from core.financials import (
    calculate_pnl,
    calculate_cashflow,
    calculate_unit_economics,
)

# --- Конфигурация страницы и стили ---
st.set_page_config(page_title="Scenario — Финансовый симулятор", layout="wide")

st.markdown(
    """
<style>
    .block-container { padding-top: 1rem !important; }
    [data-testid="stSidebar"] h2 { margin-top: -1.7rem; font-size: 24px !important; color: #262730; }
    [data-testid="stDownloadButton"] { margin-bottom: 10px; }
    div[role="radiogroup"] { flex-direction: row; gap: 2px; }
    div[role="radiogroup"] > label > div:first-child { display: none; }
    div[role="radiogroup"] > label {
        display: inline-flex; background-color: #f0f2f6; color: #555;
        padding: 8px 16px; margin: 0 !important; border-radius: 0.5rem 0.5rem 0 0;
        border: 1px solid #dcdcdc; border-bottom: 1px solid #dcdcdc;
        cursor: pointer; transition: background-color 0.2s, color 0.2s;
    }
    div[role="radiogroup"] > label:hover { background-color: #e0e2e6; }
    div[role="radiogroup"] > label:has(input:checked) {
        background-color: white; color: #333; border-bottom: 1px solid white;
    }
</style>
""",
    unsafe_allow_html=True,
)

# NEW: Надежная инициализация всех ключей session_state
for k, v in [
    ("calc_result", None),
    ("allow_send_ai", False),
    ("allow_send_ai_local", False),
    ("main_tab_selector", "pnl"),
]:
    if k not in st.session_state:
        st.session_state[k] = v


# FIXED: Убран UI-вызов из кэшируемой функции
@st.cache_data
def load_translation(lang: str = "ru") -> Dict[str, Any]:
    fname = f"{lang}.json"
    candidates = []
    try:
        here = Path(__file__).resolve().parent
        candidates.extend([here / "core" / "locales" / fname, here / "locales" / fname])
    except Exception:
        pass
    cwd = Path.cwd()
    candidates.extend([cwd / "core" / "locales" / fname, cwd / "locales" / fname])
    for p in candidates:
        try:
            if p.exists():
                with p.open("r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception:
            pass
    return {}


T = load_translation("ru")
if not T:
    st.warning("Файл локализации не найден. Будут использоваться стандартные названия.")


def _(key: str, default: str = "") -> str:
    parts = key.split(".")
    cur = T
    for p in parts:
        if isinstance(cur, dict) and p in cur:
            cur = cur[p]
        else:
            return default or key
    return str(cur)


def format_rub(value: Any, decimals: int = 0) -> str:
    try:
        if value is None:
            return "—"
        v = float(value)
        fmt = f"{v:,.{decimals}f}".replace(",", " ")
        return f"{fmt} ₽"
    except Exception:
        return f"{value} ₽"


def create_excel_workbook(
    sheets: Dict[str, pd.DataFrame], index_names: Dict[str, str]
) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            idx_name = index_names.get(sheet_name, "Index")
            df_to_write = df.copy()
            for col in df_to_write.columns:
                if pd.api.types.is_numeric_dtype(df_to_write[col]):
                    df_to_write[col] = df_to_write[col].round(2)

            df_to_write.rename_axis(idx_name).to_excel(
                writer, sheet_name=sheet_name, index=True
            )
            worksheet = writer.sheets[sheet_name]

            header_font = Font(bold=True)
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for col_idx, value in enumerate(
                df_to_write.reset_index().columns.values, 1
            ):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
            for i, column_cells in enumerate(worksheet.columns, 1):
                lengths = [
                    len(str(cell.value))
                    for cell in column_cells
                    if cell.value is not None
                ]
                max_length = max(lengths) if lengths else 0
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[get_column_letter(i)].width = adjusted_width
            for row in worksheet.iter_rows(
                min_row=2,
                max_row=worksheet.max_row,
                min_col=1,
                max_col=worksheet.max_column,
            ):
                for cell in row:
                    cell.border = border
    return output.getvalue()


# NEW: Функции для синхронизации чекбоксов
def _sync_ai_consent_local():
    st.session_state["allow_send_ai"] = st.session_state.get(
        "allow_send_ai_local", False
    )


def _sync_ai_consent_main():
    st.session_state["allow_send_ai_local"] = st.session_state.get(
        "allow_send_ai", False
    )


# --- Основной UI ---
st.title(_("app_title", "Scenario — Финансовый симулятор"))

with st.sidebar:
    st.markdown(f"## {_('inputs_header', 'Параметры бизнеса')}")

    st.checkbox(
        _("ui.ai_consent_checkbox_label", "Разрешаю отправку данных в AI"),
        key="allow_send_ai",
        on_change=_sync_ai_consent_main,
    )
    st.markdown("---")

    mrr = st.number_input(
        _("mrr_label", "MRR, ₽"), min_value=0.0, value=5000.0, step=100.0, format="%.2f"
    )
    monthly_growth_pct = st.number_input(
        _("monthly_growth_label", "Рост выручки, %"),
        min_value=0.0,
        value=5.0,
        step=0.1,
        format="%.1f",
        help="Например, 5 означает 5% роста в месяц",
    )
    churn_pct = st.number_input(
        _("churn_label", "Отток клиентов, %"),
        min_value=0.0,
        max_value=100.0,
        value=2.0,
        step=0.1,
        format="%.1f",
        help="Например, 2 означает 2% оттока в месяц",
    )
    arpu = st.number_input(
        _("arpu_label", "ARPU, ₽"), min_value=0.0, value=50.0, step=1.0, format="%.2f"
    )
    cac = st.number_input(
        _("cac_label", "CAC, ₽"), min_value=0.0, value=200.0, step=10.0, format="%.2f"
    )
    fixed_costs = st.number_input(
        _("fixed_costs_label", "Постоянные расходы, ₽"),
        min_value=0.0,
        value=3000.0,
        step=100.0,
        format="%.2f",
    )
    variable_costs_pct = st.number_input(
        _("variable_costs_label", "Переменные расходы, %"),
        min_value=0.0,
        max_value=100.0,
        value=20.0,
        step=0.1,
        format="%.1f",
    )
    payment_lag_days = st.number_input(
        _("payment_lag_label", "Задержка платежей, дней"),
        min_value=0,
        max_value=365,
        value=30,
        step=1,
    )
    horizon_months = st.number_input(
        _("horizon_label", "Горизонт, мес."),
        min_value=1,
        max_value=120,
        value=12,
        step=1,
    )

inputs = {
    "mrr": float(mrr),
    "monthly_growth_pct": float(monthly_growth_pct),
    "churn_pct": float(churn_pct),
    "arpu": float(arpu),
    "cac": float(cac),
    "fixed_costs": float(fixed_costs),
    "variable_costs_pct": float(variable_costs_pct),
    "payment_lag_days": int(payment_lag_days),
    "horizon_months": int(horizon_months),
}

INPUTS_SCHEMA = {
    "$schema": "http://json-schema.org/draft-07/schema#",
    "title": "ScenarioInputs",
    "type": "object",
    "properties": {
        "mrr": {"type": "number", "minimum": 0},
        "monthly_growth_pct": {"type": "number", "minimum": 0},
        "churn_pct": {"type": "number", "minimum": 0, "maximum": 100},
        "arpu": {"type": "number", "minimum": 0},
        "cac": {"type": "number", "minimum": 0},
        "fixed_costs": {"type": "number", "minimum": 0},
        "variable_costs_pct": {"type": "number", "minimum": 0, "maximum": 100},
        "payment_lag_days": {"type": "integer", "minimum": 0, "maximum": 365},
        "horizon_months": {"type": "integer", "minimum": 1, "maximum": 120},
    },
    "required": [
        "mrr",
        "monthly_growth_pct",
        "churn_pct",
        "arpu",
        "cac",
        "fixed_costs",
        "variable_costs_pct",
        "payment_lag_days",
        "horizon_months",
    ],
}


@st.cache_data(ttl=300)
def run_calculations(inp: Dict[str, Any]) -> Dict[str, Any]:
    local = deepcopy(inp)
    validate(instance=local, schema=INPUTS_SCHEMA)
    pnl_df = calculate_pnl(local)
    cf_df = calculate_cashflow(local)

    try:
        unit_metrics = calculate_unit_economics(local)
    except Exception:
        unit_metrics = {
            "ltv": None,
            "ltv_cac": None,
            "break_even_month": None,
            "warnings": ["UNIT_ERROR"],
        }

    return {"pnl": pnl_df, "cashflow": cf_df, "unit_metrics": unit_metrics}


try:
    with st.spinner(_("ui.calculating_spinner", "Выполняю расчеты...")):
        st.session_state.calc_result = run_calculations(inputs)
except (ValidationError, ValueError) as e:
    st.error(_("errors.invalid_input", "Ошибка: Некорректные входные данные."))
    st.json({"status": "error", "code": "INVALID_INPUT", "message": str(e)})
    st.session_state.calc_result = None
except Exception as e:
    st.error(_("errors.internal_error", "Произошла внутренняя ошибка."))
    with st.expander(_("ui.traceback_expander", "Технические детали ошибки")):
        st.exception(e)
    st.session_state.calc_result = None

if st.session_state.calc_result:
    pnl_df, cf_df, metrics = (
        st.session_state.calc_result["pnl"],
        st.session_state.calc_result["cashflow"],
        st.session_state.calc_result["unit_metrics"],
    )

    for w_code in metrics.get("warnings", []):
        st.warning(_(f"warnings.{w_code.lower()}", w_code))

    left, right = st.columns([1, 2])
    with left:
        st.header(_("unit_econ_header", "Ключевые метрики"))
        st.metric(_("ltv_label", "LTV"), format_rub(metrics.get("ltv"), decimals=0))
        st.metric(
            _("ltv_cac_label", "LTV / CAC"),
            (
                f"{metrics.get('ltv_cac'):.2f}"
                if metrics.get("ltv_cac") is not None
                else "—"
            ),
        )
        st.metric(
            _("bem_metric_label", "Месяц окупаемости"),
            (
                str(metrics.get("break_even_month"))
                if metrics.get("break_even_month") is not None
                else "—"
            ),
        )

        st.markdown("---")

        col1, col2 = st.columns([1, 3])
        with col1:
            st.checkbox(
                " ",
                key="allow_send_ai_local",
                value=st.session_state.get("allow_send_ai", False),
                on_change=_sync_ai_consent_local,
                label_visibility="hidden",
            )
        with col2:
            if st.button(_("ui.get_ai", "Получить AI-рекомендацию")):
                if not st.session_state.get("allow_send_ai", False):
                    st.info(_("ui.ai_consent_required", "Нужно разрешение"))
                else:
                    with st.spinner(_("ui.ai_spinner", "Анализирую...")):
                        st.success(_("ui.ai_stub_response", "AI: ..."))

    with right:
        st.header(_("reports_header", "Финансовые отчёты"))
        with st.expander(_("glossary.title", "Что означают эти метрики?")):
            glossary_terms = T.get("glossary", {}).get("terms", {})
            if glossary_terms:
                for term, description in glossary_terms.items():
                    st.write(f"**{term}** — {description}")
            else:
                st.write("Описание терминов не найдено.")

        headers, index_name = T.get("table_headers", {}), _(
            "table_index_month", "Месяц"
        )

        tab_options = {
            "pnl": _("pnl_tab", "P&L"),
            "cashflow": _("cashflow_tab", "Денежный поток"),
            "charts": _("charts_tab", "Графики"),
            "export": _("export_tab", "Скачать отчет"),
            "how_calc": _("how_calc_tab", "Как считается"),
        }

        selected_tab_key = st.radio(
            "tabs",
            options=list(tab_options.keys()),
            format_func=lambda key: tab_options[key],
            label_visibility="collapsed",
            horizontal=True,
            key="main_tab_selector",
        )

        st.markdown("<br>", unsafe_allow_html=True)

        if selected_tab_key == "pnl":
            st.subheader(_("pnl_subheader", "Прибыли и убытки"))
            st.dataframe(
                pnl_df.rename(columns=headers).rename_axis(index_name),
                use_container_width=True,
            )

        elif selected_tab_key == "cashflow":
            st.subheader(_("cashflow_subheader", "Движение денежных средств"))
            st.dataframe(
                cf_df.rename(columns=headers).rename_axis(index_name),
                use_container_width=True,
            )

        elif selected_tab_key == "charts":

            def create_chart(df, y_column, y_title):
                chart_data = (
                    df[[y_column]]
                    .reset_index()
                    .rename(columns={"month": index_name, y_column: y_title})
                )
                return (
                    alt.Chart(chart_data)
                    .mark_line()
                    .encode(
                        x=alt.X(f"{index_name}:Q", title=index_name),
                        y=alt.Y(
                            f"{y_title}:Q", title="Сумма", axis=alt.Axis(format="~s")
                        ),
                        tooltip=[index_name, alt.Tooltip(y_title, format=",.0f")],
                    )
                    .interactive()
                    .properties(padding={"left": 50, "bottom": 10})
                )

            st.subheader(_("charts.revenue", "Динамика выручки"))
            st.altair_chart(
                create_chart(pnl_df, "revenue", headers.get("revenue")),
                use_container_width=True,
                theme="streamlit",
            )
            st.subheader(_("charts.gross_profit", "Динамика валовой прибыли"))
            st.altair_chart(
                create_chart(pnl_df, "gross_profit", headers.get("gross_profit")),
                use_container_width=True,
                theme="streamlit",
            )
            st.subheader(_("charts.receipts_payments", "Движение денежных средств"))
            chart_data = (
                cf_df[["receipts", "payments"]]
                .reset_index()
                .rename(columns={"month": index_name, **headers})
            )
            chart_data_melted = chart_data.melt(
                id_vars=index_name, var_name="Показатель", value_name="Сумма"
            )
            chart = (
                alt.Chart(chart_data_melted)
                .mark_line()
                .encode(
                    x=alt.X(f"{index_name}:Q", title=index_name),
                    y=alt.Y("Сумма:Q", title="Сумма", axis=alt.Axis(format="~s")),
                    color=alt.Color(
                        "Показатель:N",
                        title="Показатель",
                        legend=alt.Legend(orient="right", offset=15),
                    ),
                    tooltip=[
                        index_name,
                        "Показатель",
                        alt.Tooltip("Сумма", format=",.0f"),
                    ],
                )
                .interactive()
                .properties(width=750, height=420)
            )
            st.altair_chart(chart, use_container_width=True, theme="streamlit")

        elif selected_tab_key == "export":
            param_names = T.get("parameter_names", {})
            translated_inputs = {param_names.get(k, k): v for k, v in inputs.items()}
            inputs_df = pd.DataFrame.from_dict(
                translated_inputs, orient="index", columns=["Значение"]
            )

            excel_data = create_excel_workbook(
                sheets={
                    "Параметры": inputs_df,
                    "P&L": pnl_df.rename(columns=headers),
                    "Cash Flow": cf_df.rename(columns=headers),
                },
                index_names={
                    "Параметры": "Параметр",
                    "P&L": _("table_index_month", "Месяц"),
                    "Cash Flow": _("table_index_month", "Месяц"),
                },
            )
            inputs_json = json.dumps(inputs, indent=2, ensure_ascii=False).encode(
                "utf-8"
            )

            st.download_button(
                _("ui.download_full_report", "Скачать полный отчет (Excel)"),
                excel_data,
                _("ui.full_report_filename", "финансовый_отчет.xlsx"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            st.download_button(
                _("ui.download_scenario_json", "Скачать сценарий (JSON)"),
                inputs_json,
                _("ui.scenario_filename", "сценарий.json"),
                "application/json",
            )

        elif selected_tab_key == "how_calc":
            st.subheader(_("how_calc_tab", "Как считается"))
            st.latex(r"\mathrm{LTV} = \frac{ARPU}{Churn\_rate / 100}")
            st.write(_("formulas_description", "- Выручка моделируется..."))
